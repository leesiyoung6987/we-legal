"""
저축은행 DB 변환 (convert_savings_db.py)
──────────────────────────────────────
엑셀(저축은행db정리.xlsx) → config/savings_banks.json 자동 변환

사용법: 
  1. 저축은행db정리.xlsx 수정
  2. python convert_savings_db.py 실행
"""

import json
import openpyxl
from pathlib import Path

EXCEL_PATH = Path("config/저축은행db정리.xlsx")
JSON_PATH = Path("config/savings_banks.json")


def convert():
    if not EXCEL_PATH.exists():
        print(f"❌ 엑셀 파일 없음: {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["저축은행"]

    result = {
        "_comment": "저축은행 DB. 엑셀 수정 후 python convert_savings_db.py 실행하면 자동 갱신."
    }

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        name = str(row[0]).strip() if row[0] else ""
        tel = str(row[1]).strip() if row[1] else ""
        fax = str(row[2]).strip() if row[2] else ""
        branch = str(row[3]).strip() if row[3] else ""
        if not name:
            continue

        # 중복 이름: 빈 값만 보충
        if name in result and name != "_comment":
            existing = result[name]
            if not existing.get("fax") and fax:
                existing["fax"] = fax
            if not existing.get("branch") and branch:
                existing["branch"] = branch
            continue

        result[name] = {
            "tel": tel,
            "fax": fax,
            "branch": branch
        }

    count = len(result) - 1
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    print(f"✅ 변환 완료: {count}개 저축은행 → {JSON_PATH}")


if __name__ == "__main__":
    convert()
