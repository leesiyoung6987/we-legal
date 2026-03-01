"""
debt_list_builder.py - 채권목록 엑셀 시트 생성
파싱된 신용조회 데이터를 채권목록 양식으로 엑셀에 작성합니다.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── 스타일 정의 ──
FONT_NORMAL = Font(name='맑은 고딕', size=11)
FONT_BOLD = Font(name='맑은 고딕', size=11, bold=True)
FONT_TITLE = Font(name='맑은 고딕', size=14, bold=True)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

THIN_BORDER = Border(
    left=Side('thin'), right=Side('thin'),
    top=Side('thin'), bottom=Side('thin')
)

FILLS = {
    "header": PatternFill('solid', fgColor='D9E1F2'),
    "priority": PatternFill('solid', fgColor='FFF2CC'),
    "secured": PatternFill('solid', fgColor='F8CBAD'),
    "credit": PatternFill('solid', fgColor='DAEEF3'),
    "case": PatternFill('solid', fgColor='E2EFDA'),
    "self": PatternFill('solid', fgColor='FCE4D6'),
    "nodebt": PatternFill('solid', fgColor='F2F2F2'),
}

COL_WIDTHS = {
    'B': 14.0, 'C': 6.0, 'D': 22.0, 'E': 30.0,
    'F': 14.0, 'G': 14.0, 'H': 10.0, 'I': 50.0
}


def build_debt_list_sheet(ws, classified_data):
    """워크시트에 채권목록 데이터 작성
    
    Args:
        ws: openpyxl Worksheet
        classified_data: classify_and_merge() 반환값
            {name, secured, unsecured, cards, no_debt}
    """
    # 열 너비 설정
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    row = 1

    # ── Row 1: 이름 ──
    ws.merge_cells('B1:I1')
    ws['B1'] = classified_data["name"]
    ws['B1'].font = FONT_TITLE
    ws['B1'].alignment = ALIGN_CENTER

    # ── Row 2: 헤더 ──
    headers = ['종류', '순번', '채권자명', '내역·사유', '발생일자', '특이사항', '발급여부', '조회방법']
    for i, h in enumerate(headers):
        cell = ws.cell(row=2, column=i+2, value=h)
        cell.font = FONT_BOLD
        cell.fill = FILLS["header"]
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    row = 3

    # ── 우선채권 ──
    priority = [
        {"채권자명": "국세", "내역사유": "", "발생일자": "", "조회방법": "우선채권"},
        {"채권자명": "지방세", "내역사유": "", "발생일자": "", "조회방법": "우선채권"},
        {"채권자명": "4대보험", "내역사유": "", "발생일자": "", "조회방법": "우선채권"},
    ]
    row = _write_section(ws, row, "우선채권", FILLS["priority"], priority)

    # ── 담보대출 ──
    secured = classified_data.get("secured", [])
    if not secured:
        secured_display = [{"채권자명": "", "내역사유": "", "발생일자": "", "조회방법": ""}]
    else:
        secured_display = secured
    row = _write_section(ws, row, "담보대출", FILLS["secured"], secured_display)

    # ── 채권자목록(신용조회) = 비담보 + 카드 ──
    unsecured = classified_data.get("unsecured", [])
    cards = classified_data.get("cards", [])
    credit_items = unsecured + cards
    if not credit_items:
        credit_items = [{"채권자명": "", "내역사유": "", "발생일자": "", "조회방법": ""}]
    row = _write_section(ws, row, "채권자목록\n(신용조회)", FILLS["credit"], credit_items)

    # ── 채권자목록(사건조회) - 빈칸 ──
    row = _write_empty_section(ws, row, "채권자목록\n(사건조회)", FILLS["case"], 5, "사건조회")

    # ── 채권자목록(본인진술) - 빈칸 ──
    row = _write_empty_section(ws, row, "채권자목록\n(본인진술)", FILLS["self"], 3, "본인진술")

    # ── 채무없음 ──
    no_debt = classified_data.get("no_debt", [])
    if not no_debt:
        row = _write_empty_section(ws, row, "채무없음", FILLS["nodebt"], 3, "")
    else:
        row = _write_section(ws, row, "채무없음", FILLS["nodebt"], no_debt)

    return ws


def build_debt_list_workbook(classified_data):
    """독립 엑셀 파일로 채권목록 생성
    
    Returns:
        openpyxl.Workbook
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '채권목록'
    build_debt_list_sheet(ws, classified_data)
    return wb


def update_submission_list(template_path, classified_data, output_path=None):
    """자료제출목록 엑셀의 채권목록 시트를 업데이트
    
    Args:
        template_path: 자료제출목록 엑셀 경로
        classified_data: classify_and_merge() 반환값
        output_path: 저장 경로 (None이면 template_path에 덮어쓰기)
    
    Returns:
        저장된 파일 경로
    """
    wb = openpyxl.load_workbook(template_path)
    
    # 기존 채권목록 시트 삭제 후 새로 생성
    if '채권목록' in wb.sheetnames:
        idx = wb.sheetnames.index('채권목록')
        del wb['채권목록']
        ws = wb.create_sheet('채권목록', idx)
    else:
        ws = wb.create_sheet('채권목록')
    
    build_debt_list_sheet(ws, classified_data)
    
    save_path = output_path or template_path
    wb.save(save_path)
    return save_path


# ── 내부 함수 ──

def _write_section(ws, start_row, label, fill, items):
    """데이터가 있는 섹션 작성"""
    cnt = len(items)
    if cnt > 1:
        ws.merge_cells(f'B{start_row}:B{start_row+cnt-1}')
    
    cell = ws.cell(row=start_row, column=2, value=label)
    cell.font = FONT_BOLD
    cell.fill = fill
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER

    for i, item in enumerate(items):
        r = start_row + i
        vals = [
            None,  # B열 (merge 처리됨)
            i + 1,  # 순번
            item.get("채권자명", ""),
            item.get("내역사유", ""),
            item.get("발생일자", ""),
            item.get("특이사항", ""),
            item.get("발급여부", ""),
            item.get("조회방법", ""),
        ]
        for j, v in enumerate(vals):
            col = j + 2  # B=2부터
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_LEFT if j == 7 else ALIGN_CENTER
            cell.border = THIN_BORDER
        
        # B열 병합셀 테두리
        ws.cell(row=r, column=2).border = THIN_BORDER

    return start_row + cnt


def _write_empty_section(ws, start_row, label, fill, count, col_i_text=""):
    """빈 섹션 작성 (사건조회, 본인진술, 채무없음)"""
    if count > 1:
        ws.merge_cells(f'B{start_row}:B{start_row+count-1}')
    
    cell = ws.cell(row=start_row, column=2, value=label)
    cell.font = FONT_BOLD
    cell.fill = fill
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER

    for i in range(count):
        r = start_row + i
        ws.cell(row=r, column=3, value=i+1).font = FONT_NORMAL
        ws.cell(row=r, column=3).alignment = ALIGN_CENTER
        if col_i_text:
            ws.cell(row=r, column=9, value=col_i_text).font = FONT_NORMAL
            ws.cell(row=r, column=9).alignment = ALIGN_LEFT
        for c in range(2, 10):
            ws.cell(row=r, column=c).border = THIN_BORDER

    return start_row + count
